import re
import time
from datetime import datetime
from openpyxl import load_workbook, workbook
import csv
from labcheckin.models import Student
from labcheckin import db


# returns either a swipenumber as a string or None
def parse_card(raw_card_input: str):
    '''
    https://docs.python.org/3/library/re.html#match-objects
    re.search() searches the entire string for the pattern
        returns a Match() object
    Match.group(0) returns the entire match as a single string
    '''
    match = re.search(
        r"(?<=;1234567)\d{9}(?=\=123456789012345\?)", raw_card_input)
    if match:
        return match.group(0)


# i found this on stack overflow and forgot to copy the link oops
def utc2local(utc: datetime):
    epoch = time.mktime(utc.timetuple())
    offset = datetime.fromtimestamp(epoch) - datetime.utcfromtimestamp(epoch)
    return utc + offset


def open_excel(filename):
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            print(cell.value)


def import_students_from_csv(filename):
    with open(filename, "r", encoding='ISO-8859-1') as csvfile:
        # headers = ['EKUID', 'SWIPE_NUMBER', 'FIRST_NAME', 'MIDDLE_NAME', 'LAST_NAME', 'EMAIL', 'CLASS']

        reader = csv.DictReader(csvfile, delimiter=",")
        for row in reader:
            existing_student = Student.query.filter_by(
                student_id=row["EKUID"]).first()

            if existing_student:
                def hasNumbers(inputString):
                    return any(char.isdigit() for char in inputString)

                emails = [row["EMAIL"], existing_student.email]
                correct_email = list(filter(hasNumbers, emails))[0]
                existing_student.email = correct_email
                db.session.commit()

                continue

            new_student = Student(
                student_id=row["EKUID"],
                full_name=f"{row['FIRST_NAME']} {row['MIDDLE_NAME'] + ' ' if row['MIDDLE_NAME'] else ''}{row['LAST_NAME']}",
                swipe_number=row['SWIPE_NUMBER'] if row['SWIPE_NUMBER'] else row["EKUID"],
                email=row["EMAIL"]
            )
            db.session.add(new_student)
    db.session.commit()
